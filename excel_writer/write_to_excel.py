from openpyxl import Workbook

def write_plate_info_to_excel(plate_info, excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Plate Information"

    # Write headers
    headers = ["Plate Name", "Count", "Layer Name"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Write data
    for row, plate in enumerate(plate_info, start=2):
        ws.cell(row=row, column=1, value=plate['Plate Name'])
        ws.cell(row=row, column=2, value=plate['Count'])
        ws.cell(row=row, column=3, value=plate['Layer Name'])

    wb.save(excel_path)
